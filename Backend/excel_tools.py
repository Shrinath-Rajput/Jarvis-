# excel_tools.py
"""
Excel and spreadsheet operations for JARVIS
"""

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, BarChart, Reference


class ExcelTools:
    """Handle Excel operations"""
    
    @staticmethod
    def create_spreadsheet(data, sheet_name="Sheet1", save_path=None):
        """Create a new Excel spreadsheet"""
        try:
            if save_path is None:
                save_path = os.path.expanduser("~/Documents/spreadsheet.xlsx")
            
            save_path = os.path.expanduser(save_path)
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            df = pd.DataFrame(data)
            df.to_excel(save_path, sheet_name=sheet_name, index=False)
            
            return {"success": True, "message": f"Spreadsheet created at {save_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @staticmethod
    def add_chart(excel_path, chart_type="bar", data_range=None):
        """Add a chart to Excel"""
        try:
            excel_path = os.path.expanduser(excel_path)
            
            if not os.path.exists(excel_path):
                return {"success": False, "error": f"File not found: {excel_path}"}
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            if chart_type.lower() == "pie":
                chart = PieChart()
            else:
                chart = BarChart()
            
            if data_range:
                values = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
                chart.add_data(values)
            
            ws.add_chart(chart, "D2")
            
            wb.save(excel_path)
            
            return {"success": True, "message": f"Chart added to {excel_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @staticmethod
    def import_csv(csv_path, save_path=None):
        """Import CSV to Excel"""
        try:
            csv_path = os.path.expanduser(csv_path)
            
            if not os.path.exists(csv_path):
                return {"success": False, "error": f"CSV not found: {csv_path}"}
            
            if save_path is None:
                save_path = csv_path.replace('.csv', '.xlsx')
            
            save_path = os.path.expanduser(save_path)
            
            df = pd.read_csv(csv_path)
            df.to_excel(save_path, index=False)
            
            return {"success": True, "message": f"CSV imported to {save_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @staticmethod
    def create_pivot_table(excel_path, values, index, aggfunc="sum"):
        """Create a pivot table"""
        try:
            excel_path = os.path.expanduser(excel_path)
            
            if not os.path.exists(excel_path):
                return {"success": False, "error": f"File not found: {excel_path}"}
            
            df = pd.read_excel(excel_path)
            pivot = pd.pivot_table(df, values=values, index=index, aggfunc=aggfunc)
            
            pivot_path = excel_path.replace('.xlsx', '_pivot.xlsx')
            pivot.to_excel(pivot_path)
            
            return {"success": True, "message": f"Pivot table created at {pivot_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @staticmethod
    def create_budget_tracker(categories, amounts, save_path=None):
        """Create a budget tracker"""
        try:
            if save_path is None:
                save_path = os.path.expanduser("~/Documents/budget.xlsx")
            
            save_path = os.path.expanduser(save_path)
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Budget"
            
            # Headers
            ws['A1'] = "Category"
            ws['B1'] = "Amount"
            ws['C1'] = "Percentage"
            
            total = sum(amounts)
            
            # Data
            for i, (category, amount) in enumerate(zip(categories, amounts), start=2):
                ws[f'A{i}'] = category
                ws[f'B{i}'] = amount
                ws[f'C{i}'] = f"={(amount/total)*100:.1f}%"
            
            # Total
            ws[f'A{len(categories)+2}'] = "Total"
            ws[f'B{len(categories)+2}'] = total
            
            # Add chart
            chart = PieChart()
            values = Reference(ws, min_col=2, min_row=1, max_row=len(categories)+1)
            labels = Reference(ws, min_col=1, min_row=2, max_row=len(categories)+1)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)
            
            ws.add_chart(chart, "E2")
            
            wb.save(save_path)
            
            return {"success": True, "message": f"Budget tracker created at {save_path}"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @staticmethod
    def add_formula(excel_path, cell, formula):
        """Add formula to Excel cell"""
        try:
            excel_path = os.path.expanduser(excel_path)
            
            if not os.path.exists(excel_path):
                return {"success": False, "error": f"File not found: {excel_path}"}
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            ws[cell] = formula
            
            wb.save(excel_path)
            
            return {"success": True, "message": f"Formula added to {cell}"}
        except Exception as e:
            return {"success": False, "error": str(e)}


# Export
excel_tools = ExcelTools()
